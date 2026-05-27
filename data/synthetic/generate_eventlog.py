"""
synthetic/generate_eventlog.py
==============================
KAMP event-log(LOT 이벤트·양/불 판정) 데이터의 "구조"를 모사한 합성 더미.

⚠️ 내용은 가짜지만 형태(멀티시트·라벨·불균형·NaN메타)는 실제 KAMP와 동일.
   → EVENTLOG_DATA_ROOT 한 줄로 실데이터 교체 가능.

모사하는 실제 구조 (docs/archive/data_summary.txt 기준):
  - ★멀티시트 Excel 통합 (챌린지 8): mct_tool_improve = Excel 8개 다른 시트·구조
  - ★PASS_YN 2.85% 극심 불균형 (챌린지 6): press_forming
  - ★LOT 첫 행 NaN 메타 (챌린지 3): ict_checker LOT 첫 행 Step -2/-1
  - LOT 단위 이벤트: LOT No., TimeStamp, EQP_ID, INJECTION VELOCITY...
  - CSV/Excel 혼재 (실데이터 형태)

생성 데이터셋 (data/synthetic/event-log/ 아래):
  1) press_forming.csv          — PASS_YN 2.85% 극심 불균형 (CSV) [챌린지: 극심불균형]
  2) mct_tool_improve.xlsx      — 멀티시트 3개, 각각 다른 컬럼 구조 (Excel) [챌린지: 멀티시트 통합]
  3) injection_lot.csv          — LOT 첫 행 NaN 메타 (CSV) [챌린지: LOT메타 NaN]
"""
from __future__ import annotations
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook

np.random.seed(42)
OUT = os.path.join(os.path.dirname(__file__), "event-log")


def gen_press_forming() -> dict:
    """1) PASS_YN 2.85% 극심 불균형 (CSV). ★balance_classes 트리거★."""
    n = 3000
    n_fail = int(n * 0.0285)  # 2.85% 불량
    pass_yn = np.array(["PASS"] * (n - n_fail) + ["FAIL"] * n_fail)
    np.random.shuffle(pass_yn)
    df = pd.DataFrame({
        "ITEM_CODE": [f"IT{np.random.randint(1000,9999)}" for _ in range(n)],
        "LOT_NO": [f"LOT{2024000+i//50}" for i in range(n)],
        "PRESS_FORCE": np.round(np.random.normal(120, 15, n), 2),
        "TEMP": np.round(np.random.normal(85, 5, n), 2),
        "CYCLE_TIME": np.round(np.random.normal(12, 2, n), 2),
        "PASS_YN": pass_yn,
    })
    path = os.path.join(OUT, "press_forming.csv")
    df.to_csv(path, index=False, encoding="utf-8-sig")
    ratio = (pass_yn == "FAIL").mean()
    return {"dataset": "press_forming.csv", "rows": n,
            "challenges": [f"PASS_YN 극심 불균형 (FAIL {ratio*100:.2f}%)"]}


def gen_mct_tool_improve() -> dict:
    """2) 멀티시트 Excel — 시트 3개가 각각 다른 컬럼 구조 (★멀티시트 통합 챌린지★)."""
    wb = Workbook()
    # 시트1: 온도 데이터
    ws1 = wb.active
    ws1.title = "온도측정"
    ws1.append(["LOT_NO", "TimeStamp", "EQP_ID", "TEMP_1", "TEMP_2"])
    for i in range(100):
        ws1.append([f"LOT{2024000+i}", f"2024-01-{(i%28)+1:02d} 10:{i%60:02d}",
                    f"EQP{i%5}", round(np.random.normal(85, 5), 2), round(np.random.normal(90, 5), 2)])
    # 시트2: 압력 데이터 (다른 컬럼 구조!)
    ws2 = wb.create_sheet("압력측정")
    ws2.append(["LOT_NO", "PRESSURE", "VALVE_STATE", "OPERATOR"])  # 구조 다름
    for i in range(100):
        ws2.append([f"LOT{2024000+i}", round(np.random.normal(120, 10), 2),
                    np.random.choice(["OPEN", "CLOSE"]), f"OP{i%8}"])
    # 시트3: 품질 판정 (또 다른 구조 + 라벨)
    ws3 = wb.create_sheet("품질판정")
    ws3.append(["LOT_NO", "INSPECTOR", "DEFECT_TYPE", "JUDGE"])
    for i in range(100):
        ws3.append([f"LOT{2024000+i}", f"INS{i%4}",
                    np.random.choice(["NONE", "CRACK", "SCRATCH", "NONE", "NONE"]),
                    np.random.choice(["OK", "NG"], p=[0.9, 0.1])])
    path = os.path.join(OUT, "mct_tool_improve.xlsx")
    wb.save(path)
    return {"dataset": "mct_tool_improve.xlsx", "sheets": 3,
            "challenges": ["멀티시트 3개 (온도측정/압력측정/품질판정)", "시트마다 컬럼 구조 다름"]}


def gen_injection_lot() -> dict:
    """3) LOT 첫 행 NaN 메타 (CSV). ★LOT 시작행에 Step -2/-1 메타가 NaN★."""
    rows = []
    for lot in range(20):
        lot_no = f"LOT{2024100+lot}"
        # LOT 첫 2행 = 메타 (측정값 NaN) — 실제 KAMP의 Step -2/-1 패턴
        rows.append({"LOT_NO": lot_no, "STEP": -2, "TimeStamp": "2024-01-08 11:20",
                     "EQP_ID": f"EQP{lot%5}", "INJECTION_VELOCITY": np.nan, "PRESSURE": np.nan})
        rows.append({"LOT_NO": lot_no, "STEP": -1, "TimeStamp": "2024-01-08 11:20",
                     "EQP_ID": f"EQP{lot%5}", "INJECTION_VELOCITY": np.nan, "PRESSURE": np.nan})
        # 실제 측정 행
        for step in range(1, 11):
            rows.append({"LOT_NO": lot_no, "STEP": step, "TimeStamp": "2024-01-08 11:20",
                         "EQP_ID": f"EQP{lot%5}",
                         "INJECTION_VELOCITY": round(np.random.normal(50, 8), 2),
                         "PRESSURE": round(np.random.normal(100, 12), 2)})
    df = pd.DataFrame(rows)
    path = os.path.join(OUT, "injection_lot.csv")
    df.to_csv(path, index=False, encoding="utf-8-sig")
    nan_rows = int(df["INJECTION_VELOCITY"].isna().sum())
    return {"dataset": "injection_lot.csv", "rows": len(df),
            "challenges": [f"LOT 첫 행 NaN 메타 ({nan_rows}행 Step -2/-1)"]}


def main() -> None:
    os.makedirs(OUT, exist_ok=True)
    results = [gen_press_forming(), gen_mct_tool_improve(), gen_injection_lot()]
    print("=" * 64)
    print("event-log 합성 더미 생성 완료 (KAMP 구조 모사)")
    print("=" * 64)
    for r in results:
        meta = r.get("rows", r.get("sheets"))
        print(f"\n  [{r['dataset']}] {meta}")
        for c in r["challenges"]:
            print(f"     - 챌린지: {c}")
    print(f"\n위치: {OUT}")


if __name__ == "__main__":
    main()
